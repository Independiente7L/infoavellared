import pandas as pd
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from bs4 import BeautifulSoup
import time
from datetime import datetime
import logging
import sys
import os
import random
import traceback
import shutil

# Suppress browser driver messages
os.environ['WDM_LOG_LEVEL'] = '0'

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler("transfermarkt_scraper.log"),
        logging.StreamHandler(sys.stdout)
    ]
)

# Lista completa de equipos
equipos = {
    "Arsenal FC": "https://www.transfermarkt.com.ar/arsenal-futbol-club/startseite/verein/4673",
    "AD Union Magdalena": "https://www.transfermarkt.com.ar/ad-union-magdalena/startseite/verein/14680",
    "CA Gimnasia y Esgrima (Mendoza)": "https://www.transfermarkt.com.ar/gimnasia-y-esgrima-de-mendoza/startseite/verein/14687",
    "Club Puebla": "https://www.transfermarkt.com.ar/puebla-fc/startseite/verein/5662",
    "CA Tigre": "https://www.transfermarkt.com.ar/club-atletico-tigre/startseite/verein/11831",
    "CA Chacarita Juniors": "https://www.transfermarkt.com.ar/ca-chacarita-juniors/startseite/verein/2154",
    "CA Los Andes": "https://www.transfermarkt.com.ar/ca-los-andes/startseite/verein/5208",
    "Volos NPS": "https://www.transfermarkt.com.ar/volos-nps/startseite/verein/60949",
    "CA Boston River": "https://www.transfermarkt.com.ar/ca-boston-river/startseite/verein/18074",
    "Club Deportivo Maldonado": "https://www.transfermarkt.com.ar/cd-maldonado/startseite/verein/18075",
    "Círculo Deportivo": "https://www.transfermarkt.com.ar/circulo-deportivo/startseite/verein/78072",
    "CA Platense": "https://www.transfermarkt.com.ar/club-atletico-platense/startseite/verein/928",
    "CA Nueva Chicago": "https://www.transfermarkt.com.ar/nueva-chicago/startseite/verein/10534",
    "Real Pilar FC": "https://www.transfermarkt.com.ar/real-pilar/startseite/verein/75858",
    "CSD Tristán Suárez": "https://www.transfermarkt.com.ar/csd-tristan-suarez/startseite/verein/1771",
    "CA Barracas Central": "https://www.transfermarkt.com.ar/club-atletico-barracas-central/startseite/verein/25184",
    "Barcelona SC Guayaquil": "https://www.transfermarkt.com.ar/barcelona-sc-guayaquil/startseite/verein/3523",
    "CS Dock Sud": "https://www.transfermarkt.com.ar/club-sportivo-dock-sud/startseite/verein/14675",
    "Club Atlético Tucumán": "https://www.transfermarkt.com.ar/club-atletico-tucuman/startseite/verein/14554",
    "Liverpool FC Montevideo": "https://www.transfermarkt.com.ar/liverpool-fc-montevideo/startseite/verein/10663",
    "CA Rosario Central": "https://www.transfermarkt.com.ar/club-atletico-rosario-central/startseite/verein/1418",
    "CA San Martín (San Juan)": "https://www.transfermarkt.com.ar/club-atletico-san-martin-sj-/startseite/verein/10511"
}

def crear_driver(headless=False):
    """Configura y retorna una instancia de Chrome WebDriver.
    NOTA: Es recomendado usar headless=False para asegurar que todo el contenido se cargue correctamente."""
    options = Options()
    if headless:
        options.add_argument("--headless=new")
    
    # Suppress console logging
    options.add_argument("--log-level=3")
    options.add_argument("--disable-logging")
    options.add_argument("--disable-gpu")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--window-size=1920,1080")
    
    # Random user agent
    user_agents = [
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/115.0.0.0 Safari/537.36",
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/116.0.0.0 Safari/537.36",
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/115.0.0.0 Safari/537.36"
    ]
    options.add_argument(f"user-agent={random.choice(user_agents)}")
    
    return webdriver.Chrome(options=options)

def scroll_to_view_matches(driver):
    """Simula el scroll por la página para asegurar que el contenido lazy-loaded se cargue."""
    print("Realizando scroll para cargar contenido...")
    try:
        # Scrollear hacia abajo gradualmente para activar carga de contenido
        total_height = driver.execute_script("return document.body.scrollHeight")
        viewport_height = driver.execute_script("return window.innerHeight")
        steps = min(10, max(3, int(total_height / viewport_height)))
        
        for i in range(1, steps + 1):
            scroll_position = i * total_height / steps
            driver.execute_script(f"window.scrollTo(0, {scroll_position});")
            print(f"  Scroll progreso: {i}/{steps}")
            time.sleep(0.7)  # Dar tiempo para que el contenido se cargue
        
        # Enfocarse en el área donde normalmente están los partidos (1/3 del camino hacia abajo)
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight/3);")
        print("  Enfocando sección de partidos...")
        time.sleep(2)
        
        # Volver al inicio
        driver.execute_script("window.scrollTo(0, 0);")
        time.sleep(1)
        
        print("✓ Scroll completado")
    except Exception as e:
        print(f"! Error durante scroll: {str(e)}")

def aceptar_cookies(driver):
    """Acepta cookies si aparece el banner."""
    try:
        WebDriverWait(driver, 5).until(
            EC.frame_to_be_available_and_switch_to_it((By.CSS_SELECTOR, "iframe[src*='consent']"))
        )
        aceptar_btn = WebDriverWait(driver, 5).until(
            EC.element_to_be_clickable((By.XPATH, '//button[contains(text(),"Aceptar y continuar")]'))
        )
        aceptar_btn.click()
        time.sleep(1)
        driver.switch_to.default_content()
        logging.info("Cookies aceptadas")
    except Exception:
        driver.switch_to.default_content()
        logging.info("No se encontró banner de cookies o ya fue aceptado")

def extraer_proximo_partido(driver, equipo_nombre):
    """Extrae información del próximo partido. Siempre devuelve 4 valores."""
    try:
        html = driver.page_source
        soup = BeautifulSoup(html, 'html.parser')
        
        print(f"Buscando partidos para {equipo_nombre}...")
        
        # Verificar si hay un contenedor de partidos
        matches_container = soup.find('tm-matches') or soup.find('section', class_='tm-matches')
        
        if not matches_container:
            logging.warning(f"No se encontró contenedor de partidos para {equipo_nombre}")
            return None, None, None, None, None
            
        # Buscar slides de partidos
        slides = matches_container.find_all('swiper-slide') or matches_container.find_all('div', class_='matches')
        
        if not slides:
            logging.warning(f"No se encontraron partidos para {equipo_nombre}")
            return None, None, None, None, None
        
        # Debug - contar cuántos slides/partidos hay
        print(f"Encontrados {len(slides)} partidos potenciales")
        
        for slide in slides:
            # Buscar elementos de resultado con diferentes selectores
            result_element = (
                slide.select_one('a[class*="result-state"][class*="fixture"]') or
                slide.select_one('.result-state--fixture') or
                slide.select_one('a[href*="spielbericht"]')
            )
            
            if not result_element:
                continue
                
            result_text = result_element.get_text().strip() if result_element else ""
            print(f"Resultado encontrado: '{result_text}'")
            
            # Verificar si es partido futuro
            if '-:-' in result_text or (result_element.get('class', []) and 'fixture' in ' '.join(result_element.get('class', []))):
                print(f"Es un partido futuro")
                
                # Extraer fecha y hora
                time_element = slide.find('div', class_='competition-time')
                fecha_hora = time_element.get_text().strip() if time_element else "Fecha no disponible"
                print(f"Fecha/hora: {fecha_hora}")
                
                # Extraer equipos
                teams = slide.find_all('div', class_='team')
                
                if len(teams) >= 2:
                    equipo_local_elem = teams[0].find('div', class_='club-name')
                    equipo_visitante_elem = teams[1].find('div', class_='club-name')
                    
                    local = equipo_local_elem.get_text().strip() if equipo_local_elem else "Equipo no disponible"
                    visitante = equipo_visitante_elem.get_text().strip() if equipo_visitante_elem else "Equipo no disponible"
                    
                    print(f"Local: {local}, Visitante: {visitante}")
                    
                    # También intentar obtener la competición
                    competition_element = slide.find('div', class_='competition-name')
                    competicion = competition_element.get_text().strip() if competition_element else "Desconocida"
                    print(f"Competición: {competicion}")
                    
                    # Determinar próximo rival con lógica mejorada
                    # Buscar cuál de los equipos coincide con el equipo que estamos buscando
                    equipo_busqueda_lower = equipo_nombre.lower()
                    local_lower = local.lower()
                    visitante_lower = visitante.lower()
                    
                    proximo_rival = None
                    
                    print(f"Analizando: '{equipo_nombre}' en partido '{local}' vs '{visitante}'")
                    
                    # Método 1: Coincidencia exacta por palabras clave
                    equipo_palabras_clave = []
                    
                    # Extraer palabras clave del nombre del equipo (ignorar palabras comunes)
                    palabras_comunes = ['club', 'atletico', 'ca', 'cf', 'fc', 'deportivo', 'de', 'del', 'la', 'los', 'las']
                    for palabra in equipo_busqueda_lower.split():
                        if len(palabra) > 2 and palabra not in palabras_comunes:
                            equipo_palabras_clave.append(palabra)
                    
                    print(f"Palabras clave extraídas: {equipo_palabras_clave}")
                    
                    # Verificar coincidencias en local
                    coincidencias_local = 0
                    for palabra in equipo_palabras_clave:
                        if palabra in local_lower:
                            coincidencias_local += 1
                    
                    # Verificar coincidencias en visitante
                    coincidencias_visitante = 0
                    for palabra in equipo_palabras_clave:
                        if palabra in visitante_lower:
                            coincidencias_visitante += 1
                    
                    print(f"Coincidencias - Local: {coincidencias_local}, Visitante: {coincidencias_visitante}")
                    
                    # Determinar el rival basado en las coincidencias
                    if coincidencias_local > coincidencias_visitante:
                        proximo_rival = visitante
                        print(f"Equipo identificado como LOCAL: {local} vs {visitante}")
                    elif coincidencias_visitante > coincidencias_local:
                        proximo_rival = local
                        print(f"Equipo identificado como VISITANTE: {local} vs {visitante}")
                    else:
                        # Método 2: Búsqueda más flexible si no hay coincidencias claras
                        print("No hay coincidencias claras, aplicando método alternativo...")
                        
                        # Buscar nombres simplificados (ej: "Tigre" en "Club atletico Tigre")
                        nombres_simplificados = []
                        if 'tigre' in equipo_busqueda_lower:
                            nombres_simplificados.append('tigre')
                        if 'chacarita' in equipo_busqueda_lower:
                            nombres_simplificados.append('chacarita')
                        if 'arsenal' in equipo_busqueda_lower:
                            nombres_simplificados.append('arsenal')
                        if 'platense' in equipo_busqueda_lower:
                            nombres_simplificados.append('platense')
                        if 'puebla' in equipo_busqueda_lower:
                            nombres_simplificados.append('puebla')
                        if 'barracas' in equipo_busqueda_lower:
                            nombres_simplificados.append('barracas')
                        if 'gimnasia' in equipo_busqueda_lower:
                            nombres_simplificados.append('gimnasia')
                        if 'rosario' in equipo_busqueda_lower:
                            nombres_simplificados.append('rosario')
                        if 'tucuman' in equipo_busqueda_lower:
                            nombres_simplificados.append('tucuman')
                        if 'liverpool' in equipo_busqueda_lower:
                            nombres_simplificados.append('liverpool')
                        if 'boston' in equipo_busqueda_lower:
                            nombres_simplificados.append('boston')
                        if 'maldonado' in equipo_busqueda_lower:
                            nombres_simplificados.append('maldonado')
                        if 'volos' in equipo_busqueda_lower:
                            nombres_simplificados.append('volos')
                        if 'andes' in equipo_busqueda_lower:
                            nombres_simplificados.append('andes')
                        if 'barcelona' in equipo_busqueda_lower:
                            nombres_simplificados.append('barcelona')
                        if 'dock' in equipo_busqueda_lower:
                            nombres_simplificados.append('dock')
                        if 'circulo' in equipo_busqueda_lower:
                            nombres_simplificados.append('circulo')
                        if 'chicago' in equipo_busqueda_lower:
                            nombres_simplificados.append('chicago')
                        if 'pilar' in equipo_busqueda_lower:
                            nombres_simplificados.append('pilar')
                        if 'tristan' in equipo_busqueda_lower or 'suarez' in equipo_busqueda_lower:
                            nombres_simplificados.extend(['tristan', 'suarez'])
                        if 'martin' in equipo_busqueda_lower and 'san' in equipo_busqueda_lower:
                            nombres_simplificados.extend(['martin', 'san'])
                        if 'union' in equipo_busqueda_lower and 'magdalena' in equipo_busqueda_lower:
                            nombres_simplificados.extend(['union', 'magdalena'])
                        
                        print(f"Nombres simplificados para buscar: {nombres_simplificados}")
                        
                        # Buscar con nombres simplificados
                        for nombre in nombres_simplificados:
                            if nombre in local_lower:
                                proximo_rival = visitante
                                print(f"Coincidencia encontrada en LOCAL con '{nombre}': {local} vs {visitante}")
                                break
                            elif nombre in visitante_lower:
                                proximo_rival = local
                                print(f"Coincidencia encontrada en VISITANTE con '{nombre}': {local} vs {visitante}")
                                break
                        
                        # Método 3: Si aún no se encontró, usar heurística de longitud de nombre
                        if not proximo_rival:
                            print("Aplicando heurística de longitud de nombres...")
                            # Calcular similitud por longitud y caracteres comunes
                            def calcular_similitud(str1, str2):
                                str1, str2 = str1.lower(), str2.lower()
                                caracteres_comunes = len(set(str1) & set(str2))
                                return caracteres_comunes / max(len(set(str1)), len(set(str2)))
                            
                            similitud_local = calcular_similitud(equipo_nombre, local)
                            similitud_visitante = calcular_similitud(equipo_nombre, visitante)
                            
                            print(f"Similitud - Local: {similitud_local:.2f}, Visitante: {similitud_visitante:.2f}")
                            
                            if similitud_local > similitud_visitante and similitud_local > 0.3:
                                proximo_rival = visitante
                                print(f"Por similitud, equipo identificado como LOCAL: {local} vs {visitante}")
                            elif similitud_visitante > similitud_local and similitud_visitante > 0.3:
                                proximo_rival = local
                                print(f"Por similitud, equipo identificado como VISITANTE: {local} vs {visitante}")
                            else:
                                # Último recurso: asumir local
                                proximo_rival = visitante
                                print(f"ÚLTIMO RECURSO: Asumiendo que es LOCAL: {local} vs {visitante}")
                    
                    # Validación final
                    if not proximo_rival:
                        proximo_rival = "Rival no determinado"
                        print(f"ERROR: No se pudo determinar el rival para {equipo_nombre}")
                    
                    print(f"RESULTADO FINAL - Próximo rival determinado: {proximo_rival}")
                    
                    return local, visitante, fecha_hora, competicion, proximo_rival
        
        logging.warning(f"No se encontró próximo partido para {equipo_nombre}")
        return None, None, None, None, None
        
    except Exception as e:
        logging.error(f"Error al extraer partido de {equipo_nombre}: {str(e)}")
        print(f"Error detallado: {traceback.format_exc()}")
        return None, None, None, None, None  # Siempre devolver 5 valores aunque haya error

def main():
    """Función principal que procesa todos los equipos."""
    print("\n==================================================")
    print(" SCRAPER DE PRÓXIMOS PARTIDOS - TRANSFERMARKT ")
    print("==================================================\n")
    
    fecha_actual = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    logging.info(f"Iniciando scraping de partidos en Transfermarkt - Fecha: {fecha_actual}")
    
    # Lista para almacenar los resultados
    resultados = []
    
    # Equipos procesados correctamente e incorrectamente
    equipos_ok = []
    equipos_error = []
    
    driver = None
    try:
        driver = crear_driver(headless=False)  # ¡IMPORTANTE! Mantener headless=False para asegurar carga correcta
        
        # Contador de equipos procesados
        total = len(equipos)
        contador = 0
        
        # Procesar cada equipo
        for nombre_equipo, url in equipos.items():
            contador += 1
            print(f"\n[{contador}/{total}] Procesando: {nombre_equipo}")
            print(f"URL: {url}")
            logging.info(f"Procesando: {nombre_equipo} ({contador}/{total})")
            
            for intento in range(3):  # Intentar hasta 3 veces en caso de error
                try:
                    print(f"Intento {intento+1}/3...")
                    driver.get(url)
                    
                    # Esperar carga básica
                    WebDriverWait(driver, 15).until(EC.presence_of_element_located((By.TAG_NAME, "body")))
                    
                    # Aceptar cookies
                    aceptar_cookies(driver)
                    
                    # IMPORTANTE: Realizar scroll para activar carga de contenido lazy-loaded
                    scroll_to_view_matches(driver)
                    
                    # Extraer datos - asegurándonos de recibir 5 valores
                    local, visitante, fecha_hora, competicion, proximo_rival = extraer_proximo_partido(driver, nombre_equipo)
                    
                    if local and visitante:
                        # Formatear el resultado y guardarlo
                        resultado = {
                            "equipo": nombre_equipo,
                            "local": local,
                            "visitante": visitante,
                            "Próximo Partido": fecha_hora,
                            "competicion": competicion if competicion else "Desconocida",
                            "Próximo Rival": proximo_rival if proximo_rival else "No determinado",
                            "fecha_consulta": fecha_actual
                        }
                        resultados.append(resultado)
                        equipos_ok.append(nombre_equipo)
                        print(f"✓ Éxito: {local} vs {visitante} - {fecha_hora} (Rival: {proximo_rival})")
                        logging.info(f"Éxito: {local} vs {visitante} - {fecha_hora} (Rival: {proximo_rival})")
                        break  # Salir del bucle de intentos si tuvo éxito
                    else:
                        print(f"× Intento {intento+1}: No se encontraron datos")
                        if intento == 2:  # Si es el último intento
                            resultados.append({
                                "equipo": nombre_equipo,
                                "local": "Error",
                                "visitante": "Error",
                                "fecha_hora": "Error",
                                "competicion": "Error",
                                "Proximo Rival": "Error",
                                "fecha_consulta": fecha_actual
                            })
                            equipos_error.append(nombre_equipo)
                        time.sleep(2)  # Esperar antes de reintentar
                
                except Exception as e:
                    print(f"× Error en intento {intento+1}: {str(e)[:100]}...")
                    logging.error(f"Error en intento {intento+1} para {nombre_equipo}: {str(e)}")
                    
                    if intento == 2:  # Si es el último intento
                        resultados.append({
                            "equipo": nombre_equipo,
                            "local": "Error",
                            "visitante": "Error",
                            "fecha_hora": "Error",
                            "competicion": "Error",
                            "Proximo Rival": "Error",
                            "fecha_consulta": fecha_actual
                        })
                        equipos_error.append(nombre_equipo)
                    
                    time.sleep(3)  # Esperar antes de reintentar
            
            # Pausa entre equipos para evitar bloqueos
            tiempo_espera = random.uniform(2, 5)
            print(f"Esperando {tiempo_espera:.1f}s antes del siguiente equipo...")
            time.sleep(tiempo_espera)
            
            # Guardar resultados parciales cada 5 equipos
            if contador % 5 == 0 or contador == total:
                try:
                    df_parcial = pd.DataFrame(resultados)
                    filename_parcial = f"proximos_partidos_parcial.csv"
                    df_parcial.to_csv(filename_parcial, index=False, encoding='utf-8-sig')
                    print(f"✓ Guardado parcial realizado ({contador}/{total} equipos)")
                except Exception as e:
                    print(f"! Error al guardar parcial: {str(e)}")
        
        # Crear un DataFrame con los resultados finales
        df = pd.DataFrame(resultados)
        
        # Guardar a CSV con timestamp (mantener esto para historial)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename_csv = f"proximos_partidos_{timestamp}.csv"
        df.to_csv(filename_csv, index=False, encoding='utf-8-sig')
        logging.info(f"Histórico guardado en {filename_csv}")
        
        # NUEVO: Siempre guardar en el mismo archivo Excel
        excel_filename = "proximos_partidos.xlsx"
        
        try:
            # Verificar si el Excel ya existe y hacer backup
            if os.path.exists(excel_filename):
                print(f"Actualizando archivo Excel existente: {excel_filename}")
                # Hacer una copia de respaldo del archivo anterior
                backup_file = f"backup_{excel_filename}"
                try:
                    shutil.copy2(excel_filename, backup_file)
                    print(f"✓ Backup creado: {backup_file}")
                except Exception as e:
                    print(f"! Error al crear backup: {str(e)}")
            else:
                print(f"Creando nuevo archivo Excel: {excel_filename}")
            
            # Guardar los nuevos datos
            df.to_excel(excel_filename, index=False, engine='openpyxl')
            print(f"✓ Datos guardados en Excel: {excel_filename}")
            logging.info(f"Resultados guardados en {excel_filename}")
            
            # Añadir hoja con historial de consultas
            try:
                from openpyxl import load_workbook
                wb = load_workbook(excel_filename)
                
                # Verificar si existe la hoja de historial
                hist_sheet_name = "Historial_Consultas"
                if hist_sheet_name not in wb.sheetnames:
                    wb.create_sheet(hist_sheet_name)
                    hist_sheet = wb[hist_sheet_name]
                    # Crear encabezados
                    hist_sheet.cell(row=1, column=1).value = "Fecha Consulta"
                    hist_sheet.cell(row=1, column=2).value = "Equipos Procesados"
                    hist_sheet.cell(row=1, column=3).value = "Equipos Exitosos"
                    hist_sheet.cell(row=1, column=4).value = "Equipos Fallidos"
                else:
                    hist_sheet = wb[hist_sheet_name]
                
                # Encontrar última fila
                last_row = 1
                while hist_sheet.cell(row=last_row, column=1).value is not None:
                    last_row += 1
                
                # Añadir nueva entrada de historial
                hist_sheet.cell(row=last_row, column=1).value = fecha_actual
                hist_sheet.cell(row=last_row, column=2).value = total
                hist_sheet.cell(row=last_row, column=3).value = len(equipos_ok)
                hist_sheet.cell(row=last_row, column=4).value = len(equipos_error)
                
                # Guardar el workbook actualizado
                wb.save(excel_filename)
                print(f"✓ Historial de consultas actualizado")
                
            except Exception as e:
                print(f"! No se pudo actualizar el historial: {str(e)}")
                
        except Exception as e:
            print(f"! Error al guardar Excel: {str(e)}")
            logging.error(f"Error al guardar Excel: {str(e)}")
        
        # Mostrar tabla de resultados
        print("\n==================================================")
        print(" RESUMEN DE PRÓXIMOS PARTIDOS ")
        print("==================================================")
        print(f"Total equipos: {total}")
        print(f"Equipos con datos: {len(equipos_ok)} ({', '.join(equipos_ok[:5]) + ('...' if len(equipos_ok) > 5 else '')})")
        print(f"Equipos sin datos: {len(equipos_error)} ({', '.join(equipos_error) if equipos_error else 'Ninguno'})")
        print("\nRESUMEN DE PRÓXIMOS PARTIDOS:")
        print(df[["equipo", "local", "visitante", "fecha_hora", "competicion"]])
        print(f"\nResultados guardados en: {excel_filename}")
        
    except Exception as e:
        logging.error(f"Error general: {str(e)}")
        print(f"! Error general: {str(e)}")
        
        # Intentar guardar los datos que se hayan recopilado hasta el momento
        if resultados:
            try:
                df_emergency = pd.DataFrame(resultados)
                emergency_file = f"proximos_partidos_emergency_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
                df_emergency.to_csv(emergency_file, index=False, encoding='utf-8-sig')
                print(f"✓ Datos parciales guardados en caso de emergencia: {emergency_file}")
            except:
                print("! No se pudieron guardar datos de emergencia")
    finally:
        if driver:
            driver.quit()
        logging.info("Proceso finalizado")
        print("\nProceso finalizado")

if __name__ == "__main__":
    main()
